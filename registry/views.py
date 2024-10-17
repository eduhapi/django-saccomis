# registry/views.py
import openpyxl
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .forms import UploadFileForm
from members.models import Member, MemberAccount, NextOfKin
import xlwt
import xlrd


@login_required
def download_member_excel_template(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="members_template.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Members'

    headers = [
        'First Name', 'Middle Name', 'Last Name', 'Member National ID', 'Registration Date',
        'Phone Number', 'Secondary Phone Number', 'Email', 'Date of Birth', 'Gender',
        'Client Type', 'Marital Status', 'Address', 'Town', 'Profession',
        'Employment Status', 'Employment Income', 'Spouse Name', 'Spouse Address'
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    wb.save(response)
    return response

@login_required
def upload_bulk_members(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            try:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    Member.objects.create(
                        first_name=row[0],
                        middle_name=row[1],
                        last_name=row[2],
                        member_national_id=row[3],
                        registration_date=row[4],
                        phone_number=row[5],
                        secondary_phone_number=row[6],
                        email=row[7],
                        date_of_birth=row[8],
                        gender=row[9],
                        client_type=row[10],
                        marital_status=row[11],
                        address=row[12],
                        town=row[13],
                        profession=row[14],
                        employment_status=row[15],
                        employment_income=row[16],
                        spouse_name=row[17],
                        spouse_address=row[18],
                    )
                messages.success(request, 'Members uploaded successfully.')
                return redirect('upload_bulk_members.html')
            except Exception as e:
                messages.error(request, f'Error uploading members: {e}')
    else:
        form = UploadFileForm()

    return render(request, 'registry/upload_bulk_members.html', {'form': form})


def download_member_account_excel_template(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="member_account_template.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('MemberAccount')

    columns = ['member_id', 'share_capital', 'savings', 'reg_fee', 'loan']
    for col_num, column_title in enumerate(columns):
        ws.write(0, col_num, column_title)

    wb.save(response)
    return response

def upload_bulk_member_accounts(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            workbook = xlrd.open_workbook(file_contents=file.read())
            sheet = workbook.sheet_by_index(0)
            
            for row_idx in range(1, sheet.nrows):
                member_id = int(sheet.cell(row_idx, 0).value)
                share_capital = float(sheet.cell(row_idx, 1).value)
                savings = float(sheet.cell(row_idx, 2).value)
                reg_fee = float(sheet.cell(row_idx, 3).value)
                mobile_wallet = float(sheet.cell(row_idx, 4).value)

                try:
                    member = Member.objects.get(id=member_id)
                    MemberAccount.objects.update_or_create(
                        member=member,
                        defaults={
                            'share_capital': share_capital,
                            'savings': savings,
                            'reg_fee': reg_fee,
                            'mobile_wallet': mobile_wallet,
                        }
                    )
                except Member.DoesNotExist:
                    messages.error(request, f'Member with ID {member_id} does not exist.')

            messages.success(request, 'Member Accounts uploaded successfully.')
            return redirect('registry/upload_bulk_member_accounts.html')
    else:
        form = UploadFileForm()
    return render(request, 'registry/upload_bulk_member_accounts.html', {'form': form})

def download_next_of_kin_excel_template(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="next_of_kin_template.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('NextOfKin')

    columns = ['member_id', 'full_name', 'relationship', 'kin_national_id', 'address', 'email', 'phone_number']
    for col_num, column_title in enumerate(columns):
        ws.write(0, col_num, column_title)

    wb.save(response)
    return response

def upload_bulk_next_of_kin(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            workbook = xlrd.open_workbook(file_contents=file.read())
            sheet = workbook.sheet_by_index(0)
            
            for row_idx in range(1, sheet.nrows):
                member_id = int(sheet.cell(row_idx, 0).value)
                full_name = sheet.cell(row_idx, 1).value
                relationship = sheet.cell(row_idx, 2).value
                kin_national_id = sheet.cell(row_idx, 3).value
                address = sheet.cell(row_idx, 4).value
                email = sheet.cell(row_idx, 5).value
                phone_number = sheet.cell(row_idx, 6).value

                try:
                    member = Member.objects.get(id=member_id)
                    NextOfKin.objects.update_or_create(
                        member=member,
                        full_name=full_name,
                        defaults={
                            'relationship': relationship,
                            'kin_national_id': kin_national_id,
                            'address': address,
                            'email': email,
                            'phone_number': phone_number,
                        }
                    )
                except Member.DoesNotExist:
                    messages.error(request, f'Member with ID {member_id} does not exist.')

            messages.success(request, 'Next of Kin uploaded successfully.')
            return redirect('registry/upload_bulk_next_of_kin.html')
    else:
        form = UploadFileForm()
    return render(request, 'registry/upload_bulk_next_of_kin.html', {'form': form})
